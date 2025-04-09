from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image

import cv2
import os
import glob
import shutil


class my_pyxl(object):
    def __init__(self):
        # # file name
        # file_name='db.xlsx'

        # # sheet name
        # sheet_name = '신조공사'

        # # work prameter
        # self.start_row = 'AV'
        # self.start_col = 10
        # self.end_row = 'PX'
        # self.end_col = 74
        # self.input_comm_index = 81
        # self.input_st_index = 82
        # comm_color_index='L3'
        # st_color_index='L5'

        # self.wb = None
        # self.ws = None
        # self.comm_color = None
        # self.st_color = None

        # # workbook
        # self.wb = load_workbook(file_name, data_only=True)
        # self.ws = self.wb[sheet_name]

        # self.comm_color = self.ws[comm_color_index].fill.start_color.index
        # self.st_color = self.ws[st_color_index].fill.start_color.index
        
        # run
        self.my_run()

        # self.my_fill_color()
        # self.cnt_bg()
        # self.save_file('TEST.xlsx')

    def my_run(self) :
        # while True:
        #     file_name = input("$ File Name : ")
        #     if file_name == 'exit' :
        #         return
            
        #     # load work book
        #     wb = load_workbook(file_name, data_only=True)
            
        #     # load sheet
        #     sheet_name = input("$ Sheet Name : ")
        #     ws = wb[sheet_name]

        #     # set cell
        #     target_cell = input("Target Cell : ")
        #     input_data = input("Data : ")
        #     target_cnt = input("Count : ")

        #     for i in range(int(target_cnt)) :
        f = input('****\n1 : Make resized image\n2 : Make IR report\n3 : Make XE4 report\n4 : Make 4K report\n5 : Make Client report\n****\n$ ')
        if f == '1' :
            self.make_resized_image()
        elif f == '2' :
            self.make_IR_gara_from_real_data()
        elif f == '3' :
            self.make_XE4_gara_from_real_data()
        elif f == '4' :
            self.make_4K_gara_from_real_data()
        elif f == '5' :
            self.make_Client_gara_from_real_data()
        elif f == '6' :
            self.make_SBC_gara_from_real_data()
        elif f == 'exit' :
            exit()
        else :
            self.my_run()


    def make_IR_gara_from_real_data(self):
        file_name = input("$ Origin File Name : ")
        
        # load work book
        wb = load_workbook(file_name, data_only=True)
        
        # load sheet
        sheet_name = input("$ Sheet Name : ")
        ws = wb[sheet_name]

        # load image path
        all_image_path = input("$ All Image Path : ")
        all_image_path = './' + all_image_path + '/'

        all_images = os.listdir(all_image_path)
        # all_images = os.listdir('./2025-1-IR/')

        # get target directory
        target_dir = input('$ Target dir : ')

        # get target file name
        target_file_name = input('$ Target File Name + {S/N} : ')

        cnt = 0
        qc_names = ['ㅁ', 'ㅎ']
        cell_qc_name = 'Z7'
        date = ['2024.09.19', '2024.09.20']
        cell_date = 'Z5'
        cell_sn = 'J7'

        for image_name in all_images :
            wb = load_workbook(file_name, data_only=True)
            ws = wb[sheet_name]
            cnt += 1
            image_path = all_image_path + image_name
            image = Image(image_path)
            ws.add_image(image, 'H10')
            ws[cell_qc_name].value = qc_names[0] # 0 : ㅁ / 1 : ㅎ
            ws[cell_sn].value = image_name[:-4]
            if cnt < 50 :
                ws[cell_date] = date[0]
            else:
                ws[cell_date] = date[1]

            save_path = target_dir + target_file_name + '_' + image_name[:-4] + '.xlsx'
            wb.save(save_path)
            print('Save : ' + save_path)
            # break


    def make_4K_gara_from_real_data(self):
        file_name = input("$ Origin File Name : ")
        
        # load work book
        wb = load_workbook(file_name, data_only=True)
        
        # load sheet
        sheet_name = input("$ Sheet Name : ")
        ws = wb[sheet_name]

        # load image path
        all_image_path = input("$ All Image Path : ")
        all_image_path = './' + all_image_path + '/'

        all_images = os.listdir(all_image_path)

        # get target directory
        target_dir = input('$ Target dir : ')

        # get target file name
        target_file_name = input('$ Target File Name + {S/N} : ')

        cnt = 0
        qc_names = ['ㅁ', 'ㅎ']
        cell_qc_name = 'Z7'
        date = ['2024.08.13', '2024.08.14', '2024.08.15', '2024.08.16', '2024.08.17']
        cell_date = 'Z5'
        cell_sn = 'J7'

        for image_name in all_images :
            wb = load_workbook(file_name, data_only=True)
            ws = wb[sheet_name]
            cnt += 1
            image_path = all_image_path + image_name
            image = Image(image_path)
            ws.add_image(image, 'H10')
            # ws[cell_qc_name].value = qc_names[0] # 0 : ㅁ / 1 : ㅎ
            ws[cell_sn].value = image_name[:-4]
            if cnt < 50 :
                ws[cell_date] = date[0]
            elif cnt < 100:
                ws[cell_date] = date[1]
            elif cnt < 150 :
                ws[cell_date] = date[2]
            elif cnt < 200 :
                ws[cell_date] = date[3]
            else :
                ws[cell_date] = date[4]

            save_path = target_dir + target_file_name + '_' + image_name[:-4] + '.xlsx'
            wb.save(save_path)
            print('Save : ' + save_path)
            # break

    
    def make_XE4_gara_from_real_data(self) :
        # 혜인님 실제 데이터
        hy_file = '원자재.xlsx'
        hy_wb = load_workbook(hy_file, data_only=True)
        hy_ws = hy_wb['XE4']

        in_date = 'C'
        out_date = 'G'
        serial_no = 'H'
        type = 'E'
        yard = 'D'
        hull_num = 'F'

        # load work book
        origin_file = './QC검사성적서/QC검사성적서 - Main Server 샘플.xlsx'
        # wb = load_workbook(origin_file, data_only=True)
        # ws = wb['검사성적서 - Main Server']
        target_dir = './gara/MAIN_SERVER/'

        # DATA
        qc_names = ['ㅁ', 'ㅎ']
        cell_input_date = 'S5'
        cell_qc_name = 'Z7'
        cell_check_date = 'Z5'
        cell_sn = 'J7'
        cell_hull_num1 = 'I32'
        cell_hull_num2 = 'M32'

        for i in range (132, 262) :
            cell_in_date = in_date + str(i)
            cell_out_date = out_date + str(i)
            cell_serial_no = serial_no + str(i)
            cell_type = type + str(i)
            cell_hull = hull_num + str(i)
            cell_yard = yard + str(i)
            

            wb = load_workbook(origin_file)
            ws = wb['검사성적서 - Main Server']
            ws[cell_qc_name] = qc_names[0]
            ws[cell_sn] = hy_ws[cell_serial_no].value
            ws[cell_input_date] = hy_ws[cell_in_date].value
            ws[cell_check_date] = hy_ws[cell_out_date].value
            ws[cell_hull_num1] = hy_ws[cell_yard].value
            ws[cell_hull_num2] = hy_ws[cell_hull].value
            
            save_path = target_dir + '검사성적서_XE4_' + hy_ws[cell_serial_no].value + '.xlsx'
            wb.save(save_path)
            print('Save : ' + save_path)


    def make_Client_gara_from_real_data(self) :
        # 혜인님 실제 데이터
        hy_file = '원자재.xlsx'
        hy_wb = load_workbook(hy_file, data_only=True)
        hy_ws = hy_wb['CLIENT(XE4)']

        in_date = 'C'
        out_date = 'G'
        serial_no = 'H'
        type = 'E'
        yard = 'D'
        hull_num = 'F'

        # load work book
        origin_file = './QC검사성적서/QC검사성적서 - Client Server 샘플.xlsx'
        # wb = load_workbook(origin_file, data_only=True)
        # ws = wb['검사성적서 - Main Server']
        target_dir = './gara/CLIENT_SERVER/'

        # DATA
        qc_names = ['ㅁ', 'ㅎ']
        cell_input_date = 'S5'
        cell_qc_name = 'Z7'
        cell_check_date = 'Z5'
        cell_sn = 'J7'
        cell_hull_num1 = 'I32'
        cell_hull_num2 = 'M32'

        for i in range (30, 48) :
            cell_in_date = in_date + str(i)
            cell_out_date = out_date + str(i)
            cell_serial_no = serial_no + str(i)
            # cell_type = type + str(i)
            cell_hull = hull_num + str(i)
            cell_yard = yard + str(i)
            

            wb = load_workbook(origin_file)
            ws = wb['검사성적서 - Client Server']
            ws[cell_qc_name] = qc_names[0]
            ws[cell_sn] = hy_ws[cell_serial_no].value
            ws[cell_input_date] = hy_ws[cell_in_date].value
            ws[cell_check_date] = hy_ws[cell_out_date].value
            ws[cell_hull_num1] = hy_ws[cell_yard].value
            ws[cell_hull_num2] = hy_ws[cell_hull].value
            
            save_path = target_dir + '검사성적서_CLIENT_' + hy_ws[cell_serial_no].value + '.xlsx'
            wb.save(save_path)
            print('Save : ' + save_path)


    def make_SBC_gara_from_real_data(self) :
        # 혜인님 실제 데이터
        hy_file = '원자재.xlsx'
        hy_wb = load_workbook(hy_file, data_only=True)
        hy_ws = hy_wb['HUBBLE']

        in_date = 'C'
        out_date = 'G'
        serial_no = 'I'
        type = 'E'
        yard = 'D'
        hull_num = 'F'

        # load work book
        origin_file = './QC검사성적서/QC검사성적서 - Single Board Computer 샘플.xlsx'
        # wb = load_workbook(origin_file, data_only=True)
        # ws = wb['검사성적서 - SingleBoardComputer']
        target_dir = './gara/SBC/'

        # DATA
        qc_names = ['ㅁ', 'ㅎ']
        cell_input_date = 'S5'
        cell_qc_name = 'Z7'
        cell_check_date = 'Z5'
        cell_sn = 'J7'
        cell_hull_num1 = 'K32'
        cell_hull_num2 = 'N32'

        for i in range (83, 187) :
            cell_in_date = in_date + str(i)
            cell_out_date = out_date + str(i)
            cell_serial_no = serial_no + str(i)
            # cell_type = type + str(i)
            cell_hull = hull_num + str(i)
            cell_yard = yard + str(i)
            

            wb = load_workbook(origin_file)
            ws = wb['검사성적서 - SingleBoardComputer']
            ws[cell_qc_name] = qc_names[i%2]
            ws[cell_sn] = hy_ws[cell_serial_no].value
            ws[cell_input_date] = hy_ws[cell_in_date].value
            ws[cell_check_date] = hy_ws[cell_out_date].value
            ws[cell_hull_num1] = hy_ws[cell_yard].value
            ws[cell_hull_num2] = hy_ws[cell_hull].value
            
            save_path = target_dir + '검사성적서_SingleBoardComputer_' + hy_ws[cell_serial_no].value + '.xlsx'
            wb.save(save_path)
            print('Save : ' + save_path)


    def make_resized_image(self):
        '''
        엑셀에 넣을 사진 크기 임의로 조정하는 함수
        '''
        all_image_path = './2025-1-4K/'
        all_images = os.listdir(all_image_path)
        cnt = 0
        for image_name in all_images :
            cnt += 1
            origin_image_path = all_image_path + image_name
            resized_image_path = './2025-1-4K-resized/' + '25' + str(format(cnt, '03')) + '.png' # + image_name

            image = cv2.imread(origin_image_path)
            resized_image = cv2.resize(image, (300,200))
            cv2.imwrite(resized_image_path, resized_image)
            print(resized_image_path)


    def next_row(self, r) :
        '''
        max row = YZ
        '''
        a,b = ord(r[0]), ord(r[1])
        
        if b == 90 :
            b = chr(65)
            a = chr(a+1)
        else :
            b = chr(b+1)
            a = chr(a)

        return a+b


    def prov_row(self, r) :
        '''
        min row = AV
        '''
        a,b = ord(r[0]), ord(r[1])

        if b == 65 :
            b = chr(90)
            a = chr(a-1)
        else :
            b = chr(b-1)
            a = chr(a)
        
        return a+b


    def get_type(self, type):
        if 'CONTROL' in type :
            return 'CON'
        if 'Contorl' in type:
            return 'CON'
        if '2.0' in type:
            return 'CON'
        if 'SVM' in type:
            return 'SVM'
        if 'DOM' in type:
            return 'DOM'
        return 'NAS'    


    def comm_fill_color(self, comm_color, start_index, product_type):
        '''
        Comm'
        NAS : 3~4 days
        DOM : 3~4 days
        SVM : 5~7 days
        CON : 5~7 days

        S/T
        HHI : 4~5 days
        HMD : 4~5 days
        HSHI : 7 days
        '''
        if product_type is None :
            return
        r, c = start_index[:2], start_index[2:]
        type_dic = {'NAS':4, 'DOM':4, 'CON':7, 'SVM':7}
        target_index = self.next_row(r) + c
        cnt = type_dic[self.get_type(product_type)]

        for i in range(cnt-1) :
            self.ws[target_index].fill = PatternFill(start_color=comm_color, end_color=comm_color, fill_type='solid')
            target_index = self.next_row(target_index[:2]) + c
    

    def st_fill_color(self, st_color, start_index, product_type):
        '''
        Comm'
        NAS : 3~4 days
        DOM : 3~4 days
        SVM : 5~7 days
        CON : 5~7 days

        S/T
        HHI : 4~5 days
        HMD : 4~5 days
        HSHI : 7 days
        '''
        if product_type is None :
            return
        r, c = start_index[:2], start_index[2:]
        type_dic = {'HHI':5, 'HMD':5, 'HSHI':7}
        target_index = self.next_row(r) + c
        cnt = type_dic[self.get_type(product_type)]

        for i in range(cnt-1) :
            self.ws[target_index].fill = PatternFill(start_color=st_color, end_color=st_color, fill_type='solid')
            target_index = self.next_row(target_index[:2]) + c


    def cnt_bg(self) :
        # count col by bg_color
        temp_row = self.start_row
        while temp_row != self.end_row :
            cur_row = temp_row
            temp_row = self.next_row(temp_row)
            comm_cnt = 0
            st_cnt = 0
            for i in range(self.start_col, self.end_col + 1) :
                cur_cell_index = cur_row + str(i)
                cur_cell_color = self.ws[cur_cell_index].fill.start_color.index
                if cur_cell_color == self.comm_color :
                    comm_cnt += 1
                if cur_cell_color == self.st_color :
                    st_cnt += 1
            comm_write_index = cur_row + str(self.input_comm_index)
            st_write_index = cur_row + str(self.input_st_index)

            self.ws[comm_write_index].value = comm_cnt
            self.ws[st_write_index].value = st_cnt
        

    def my_fill_color(self) :
        # fill_color_comm'
        temp_row = self.end_row
        while temp_row != self.start_row :
            cur_row = temp_row
            temp_row = self.prov_row(temp_row)
            for i in range(self.start_col, self.end_col + 1) :
                cur_cell_index = cur_row + str(i)
                                
                if self.ws[cur_cell_index].fill.start_color.index != self.comm_color:
                    pass
                else:
                    product_type = self.ws['E'+str(i)].value
                    self.comm_fill_color(self.comm_color, cur_cell_index, product_type)
        
        # fill_color_S/T'
        temp_row = self.end_row
        while temp_row != self.start_row :
            cur_row = temp_row
            temp_row = self.prov_row(temp_row)
            for i in range(self.start_col, self.end_col + 1) :
                cur_cell_index = cur_row + str(i)
                                
                if self.ws[cur_cell_index].fill.start_color.index != self.st_color:
                    pass
                else:
                    product_type = self.ws['F'+str(i)].value
                    self.comm_fill_color(self.st_color, cur_cell_index, product_type)


    def save_file(self, file_name):
        self.wb.save(file_name)


test = my_pyxl()
