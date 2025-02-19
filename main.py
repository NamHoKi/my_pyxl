from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class my_pyxl(object):
    def __init__(self):
        # file name
        file_name='db.xlsx'

        # sheet name
        sheet_name = '신조공사'

        # work prameter
        self.start_row = 'AV'
        self.start_col = 10
        self.end_row = 'PX'
        self.end_col = 74
        self.input_comm_index = 81
        self.input_st_index = 82
        comm_color_index='L3'
        st_color_index='L5'

        # workbook
        self.wb = load_workbook(file_name, data_only=True)
        self.ws = self.wb[sheet_name]

        self.comm_color = self.ws[comm_color_index].fill.start_color.index
        self.st_color = self.ws[st_color_index].fill.start_color.index

        self.my_fill_color()
        self.cnt_bg()
        self.save_file('TEST.xlsx')


    def next_row(self, r) :
        '''
        max row = YZ
        '''
        a,b = ord(r[0]), ord(r[1])
        
        if b == 90 :
            b = chr(65)
            a = chr(a+1)
        else :
            b = chr(b+1)
            a = chr(a)

        return a+b


    def prov_row(self, r) :
        '''
        min row = AV
        '''
        a,b = ord(r[0]), ord(r[1])

        if b == 65 :
            b = chr(90)
            a = chr(a-1)
        else :
            b = chr(b-1)
            a = chr(a)
        
        return a+b


    def get_type(self, type):
        if 'CONTROL' in type :
            return 'CON'
        if 'Contorl' in type:
            return 'CON'
        if '2.0' in type:
            return 'CON'
        if 'SVM' in type:
            return 'SVM'
        if 'DOM' in type:
            return 'DOM'
        return 'NAS'    


    def comm_fill_color(self, comm_color, start_index, product_type):
        '''
        Comm'
        NAS : 3~4 days
        DOM : 3~4 days
        SVM : 5~7 days
        CON : 5~7 days

        S/T
        HHI : 4~5 days
        HMD : 4~5 days
        HSHI : 7 days
        '''
        if product_type is None :
            return
        r, c = start_index[:2], start_index[2:]
        type_dic = {'NAS':4, 'DOM':4, 'CON':7, 'SVM':7}
        target_index = self.next_row(r) + c
        cnt = type_dic[self.get_type(product_type)]

        for i in range(cnt-1) :
            self.ws[target_index].fill = PatternFill(start_color=comm_color, end_color=comm_color, fill_type='solid')
            target_index = self.next_row(target_index[:2]) + c
    

    def st_fill_color(self, st_color, start_index, product_type):
        '''
        Comm'
        NAS : 3~4 days
        DOM : 3~4 days
        SVM : 5~7 days
        CON : 5~7 days

        S/T
        HHI : 4~5 days
        HMD : 4~5 days
        HSHI : 7 days
        '''
        if product_type is None :
            return
        r, c = start_index[:2], start_index[2:]
        type_dic = {'HHI':5, 'HMD':5, 'HSHI':7}
        target_index = self.next_row(r) + c
        cnt = type_dic[self.get_type(product_type)]

        for i in range(cnt-1) :
            self.ws[target_index].fill = PatternFill(start_color=st_color, end_color=st_color, fill_type='solid')
            target_index = self.next_row(target_index[:2]) + c


    def cnt_bg(self) :
        # count col by bg_color
        temp_row = self.start_row
        while temp_row != self.end_row :
            cur_row = temp_row
            temp_row = self.next_row(temp_row)
            comm_cnt = 0
            st_cnt = 0
            for i in range(self.start_col, self.end_col + 1) :
                cur_cell_index = cur_row + str(i)
                cur_cell_color = self.ws[cur_cell_index].fill.start_color.index
                if cur_cell_color == self.comm_color :
                    comm_cnt += 1
                if cur_cell_color == self.st_color :
                    st_cnt += 1
            comm_write_index = cur_row + str(self.input_comm_index)
            st_write_index = cur_row + str(self.input_st_index)

            self.ws[comm_write_index].value = comm_cnt
            self.ws[st_write_index].value = st_cnt
        

    def my_fill_color(self) :
        # fill_color_comm'
        temp_row = self.end_row
        while temp_row != self.start_row :
            cur_row = temp_row
            temp_row = self.prov_row(temp_row)
            for i in range(self.start_col, self.end_col + 1) :
                cur_cell_index = cur_row + str(i)
                                
                if self.ws[cur_cell_index].fill.start_color.index != self.comm_color:
                    pass
                else:
                    product_type = self.ws['E'+str(i)].value
                    self.comm_fill_color(self.comm_color, cur_cell_index, product_type)
        
        # fill_color_S/T'
        temp_row = self.end_row
        while temp_row != self.start_row :
            cur_row = temp_row
            temp_row = self.prov_row(temp_row)
            for i in range(self.start_col, self.end_col + 1) :
                cur_cell_index = cur_row + str(i)
                                
                if self.ws[cur_cell_index].fill.start_color.index != self.st_color:
                    pass
                else:
                    product_type = self.ws['F'+str(i)].value
                    self.comm_fill_color(self.st_color, cur_cell_index, product_type)


    def save_file(self, file_name):
        self.wb.save(file_name)


test = my_pyxl()
